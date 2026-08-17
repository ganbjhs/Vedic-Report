"""Upload handling: parse → validate → normalize to a canonical .xlsx.

Why normalize instead of handing the raw upload to the pipeline:

  * `src/input_loader.py` routes every path through `load_excel`, which needs a
    real .xlsx. A legacy binary `.xls` fails the zip check and is then read as
    text — producing garbage. A `.tsv` is parsed with csv's *comma* dialect, so
    "Name<TAB>URL" collapses into one cell and the link comes out as
    "Name\\thttps://…". Normalizing here fixes both **without editing the frozen
    loader** (rule #1).
  * It lets us reject a bad file with a useful message *before* a job is queued,
    and show the user how many links were found.

The layout logic itself (header row vs plain list, running category headers, URL
cleanup, X-only filtering) is NOT reimplemented — we build a grid and hand it to
the frozen `input_loader._rows_from_grid`, so the web app and the CLI agree on
exactly what a given sheet means.
"""
import csv
import io
import re
import sys
from pathlib import Path

from . import config

# The frozen loader is imported read-only — never modified.
sys.path.insert(0, str(config.ROOT / "src"))
if str(config.ROOT / "profiles") not in sys.path:
    sys.path.insert(0, str(config.ROOT / "profiles"))     # netlinks (platform links)
import input_loader  # noqa: E402

ALLOWED_SUFFIXES = (".xlsx", ".xls", ".csv", ".tsv", ".txt")

_URL_RE = re.compile(r"https?://\S+", re.I)


class UploadError(Exception):
    """A user-facing problem with the uploaded file."""


# --------------------------------------------------------------------------- #
# Name sanitizing
# --------------------------------------------------------------------------- #
def safe_stem(name: str, fallback: str = "Report") -> str:
    """Turn the user's 'Report / File Name' into a filesystem-safe stem.

    Same spirit as run.py's `re.sub(r"[^0-9A-Za-z._-]+", "_", …)`, plus explicit
    traversal defence: no separators, no leading dots, length capped.
    """
    name = (name or "").strip()
    name = name.replace("/", " ").replace("\\", " ")
    stem = re.sub(r"[^0-9A-Za-z._-]+", "_", name).strip("._-")
    stem = re.sub(r"_{2,}", "_", stem)[:60]
    return stem or fallback


def display_title(name: str, fallback: str = "Report") -> str:
    """The human title shown in the document header — punctuation kept, control
    characters and newlines removed."""
    title = re.sub(r"[\x00-\x1f\x7f]", " ", (name or "")).strip()
    title = re.sub(r"\s{2,}", " ", title)[:80]
    return title or fallback


# Characters no mainstream filesystem accepts in a name. Everything else the
# user typed — spaces, dashes, brackets, non-Latin script — is kept, because the
# download is meant to read exactly like the header of the document inside it.
_ILLEGAL_IN_FILENAME = str.maketrans({c: " " for c in '/\\:*?"<>|'})


def download_name(title: str, fallback: str = "Report") -> str:
    """The name the browser saves the file under: the user's own text.

    This is a Content-Disposition value, never a path — the file on disk keeps
    the conservative `safe_stem` name — so only characters that would break a
    save dialog are replaced. A reserved Windows device name is suffixed rather
    than mangled, since "CON.pdf" cannot be written on Windows at all.
    """
    name = display_title(title, fallback).translate(_ILLEGAL_IN_FILENAME)
    name = re.sub(r"\s{2,}", " ", name).strip(" .")
    if not name:
        return fallback
    if name.upper() in {"CON", "PRN", "AUX", "NUL"} or \
            re.fullmatch(r"(?:COM|LPT)[1-9]", name.upper()):
        name += "_"
    return name


def safe_upload_name(filename: str) -> str:
    """A display-only version of the uploaded filename. Never used as a path."""
    base = Path((filename or "").replace("\\", "/")).name
    return re.sub(r"[\x00-\x1f\x7f]", "", base)[:120] or "upload"


# --------------------------------------------------------------------------- #
# Format readers -> grid (list of rows of strings)
# --------------------------------------------------------------------------- #
def _decode(raw: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "utf-16", "cp1252", "latin-1"):
        try:
            return raw.decode(enc)
        except (UnicodeDecodeError, UnicodeError):
            continue
    return raw.decode("utf-8", errors="replace")


def _pick_sheet(names, sheet):
    """Resolve a requested tab to a name, or None for 'whatever opens'."""
    if sheet is None or sheet == "":
        return None
    if isinstance(sheet, int) or (isinstance(sheet, str) and sheet.isdigit()):
        i = int(sheet)
        if 0 <= i < len(names):
            return names[i]
        raise UploadError(f"That file has {len(names)} tab(s); tab {i} does not exist.")
    if sheet in names:
        return sheet
    raise UploadError(
        f"No tab named {sheet!r}. Available: {', '.join(names) or '(none)'}.")


def _grid_from_xlsx(path: Path, sheet=None) -> list:
    try:
        from openpyxl import load_workbook
    except ImportError as e:                                  # pragma: no cover
        raise UploadError("Server is missing openpyxl — cannot read .xlsx.") from e
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise UploadError(
            "That .xlsx could not be opened as an Excel workbook. If it is "
            "really a CSV or text file, rename it to .csv or .txt and re-upload."
        ) from e
    try:
        name = _pick_sheet(list(wb.sheetnames), sheet)
        ws = wb[name] if name else wb.active
        grid = [[("" if c is None else str(c)).strip() for c in row]
                for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    return grid


def _grid_from_xls(path: Path, sheet=None) -> list:
    try:
        import xlrd
    except ImportError as e:
        raise UploadError(
            "Server is missing xlrd — cannot read legacy .xls files. "
            "Re-save the file as .xlsx or .csv and upload again.") from e
    try:
        book = xlrd.open_workbook(str(path))
    except Exception as e:
        raise UploadError(
            "That .xls could not be opened. Re-save it as .xlsx or .csv."
        ) from e
    name = _pick_sheet(list(book.sheet_names()), sheet)
    ws = book.sheet_by_name(name) if name else book.sheet_by_index(0)
    grid = []
    for r in range(ws.nrows):
        row = []
        for c in range(ws.ncols):
            v = ws.cell_value(r, c)
            if isinstance(v, float) and v.is_integer():
                v = int(v)
            row.append(str(v).strip() if v is not None else "")
        grid.append(row)
    return grid


def _sniff_delimiter(text: str, default: str) -> str:
    sample = "\n".join(text.splitlines()[:20])
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except Exception:
        # Fall back on whichever candidate actually appears most often.
        counts = {d: sample.count(d) for d in (",", "\t", ";", "|")}
        best = max(counts, key=counts.get)
        return best if counts[best] else default


def _grid_from_text(path: Path, default_delim: str) -> list:
    text = _decode(path.read_bytes())
    if not text.strip():
        raise UploadError("That file is empty.")
    delim = _sniff_delimiter(text, default_delim)
    if not isinstance(delim, str) or len(delim) != 1:
        delim = ","
    try:
        rows = list(csv.reader(io.StringIO(text), delimiter=delim))
    except csv.Error:
        rows = [[line] for line in text.splitlines()]
    return [[(c or "").strip() for c in row] for row in rows]


def list_sheets(path: Path, suffix: str) -> list:
    """Tab names in a workbook, in order. Empty for non-workbook formats.

    `load_workbook` silently uses `wb.active`, so a file whose links live on the
    second tab used to produce "no links found" with no hint that other tabs
    existed. Listing them is what lets the user say which one.
    """
    suffix = (suffix or path.suffix).lower()
    try:
        if suffix == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            try:
                return list(wb.sheetnames)
            finally:
                wb.close()
        if suffix == ".xls":
            import xlrd
            return list(xlrd.open_workbook(str(path)).sheet_names())
    except Exception:
        return []
    return []


def read_grid(path: Path, suffix: str, sheet=None) -> list:
    """`sheet` selects a workbook tab by NAME (or index); None keeps the
    previous behaviour of using whichever tab the file opens on."""
    suffix = (suffix or path.suffix).lower()
    if suffix == ".xlsx":
        return _grid_from_xlsx(path, sheet)
    if suffix == ".xls":
        return _grid_from_xls(path, sheet)
    if suffix == ".tsv":
        return _grid_from_text(path, "\t")
    if suffix in (".csv", ".txt"):
        # A one-link-per-line .txt parses cleanly as single-column CSV, which is
        # exactly how the frozen loader treats a pasted list.
        return _grid_from_text(path, ",")
    raise UploadError(
        f"Unsupported file type '{suffix}'. Accepted: "
        + ", ".join(ALLOWED_SUFFIXES) + ".")


# --------------------------------------------------------------------------- #
# Parse + validate + write the canonical workbook
# --------------------------------------------------------------------------- #
def parse_rows(path: Path, suffix: str) -> list:
    """Uploaded file -> the pipeline's row contract, using the frozen loader's
    own layout logic. Raises UploadError with a message fit for the UI."""
    grid = read_grid(path, suffix)
    if not any(any(c for c in row) for row in grid):
        raise UploadError("That file has no content.")

    total_urls = sum(1 for row in grid for c in row if _URL_RE.search(c or ""))
    rows = input_loader._rows_from_grid(grid)   # read-only reuse — never edited

    if not rows:
        if total_urls:
            raise UploadError(
                f"Found {total_urls} link(s), but none of them are X/Twitter "
                "post links. This tool only captures x.com / twitter.com posts.")
        raise UploadError(
            "No links found. Put one X/Twitter post URL per row (or per line), "
            "or use a sheet with a column headed 'Link'.")
    if len(rows) > config.MAX_LINKS:
        raise UploadError(
            f"That file has {len(rows)} X links — the limit is "
            f"{config.MAX_LINKS} per job. Split it into smaller files.")
    return rows


def write_canonical_xlsx(rows: list, dest: Path) -> None:
    """Write rows as an Account | Link sheet.

    Those exact headers are recognised by `input_loader._header_index`, so the
    pipeline reads back precisely the rows we validated, in the same order.

    NOTE — the category column is deliberately omitted. A sheet's category is
    almost always a section word like "Tweet Links", and the Twitter report
    prints it above every screenshot, which is just noise in the finished
    document. `input_loader` defaults a missing category to "Uncategorized", and
    the report builder only prints the label when some row has a real category —
    so leaving the column out removes it from the document without touching the
    frozen builder. (The Influencer report shows the account name instead.)
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Links"
    # A combined report keeps its sections and the sheet's metric columns; the
    # X-only reports keep the two-column shape (their builders print the
    # category above every screenshot, which reads as noise — see below).
    keys = []
    for r in rows:
        for k in (r.get("sheet_metrics") or {}):
            if k not in keys:
                keys.append(k)
    with_sections = any((r.get("platform") not in (None, "x")) and
                        (r.get("category") not in (None, "", "Uncategorized")) for r in rows)
    header = ["Account", "Link"] + (["Section"] if with_sections else []) + [k.title() for k in keys]
    ws.append(header)
    for r in rows:
        # An `account_auto` row had NO name in the source sheet — what it
        # carries is a placeholder derived from the URL. Writing that placeholder
        # here would read back as a real name, and the worker would then keep
        # "Facebook post" instead of the page name the capture reads off the
        # post. Leaving the cell empty round-trips the flag through the
        # canonical sheet, and the reader derives the same placeholder again.
        account = "" if r.get("account_auto") else r.get("account_name", "")
        row = [account, r.get("link", "")]
        if with_sections:
            row.append(r.get("category") or "")
        m = r.get("sheet_metrics") or {}
        row += [m.get(k, "") for k in keys]
        ws.append(row)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


# --------------------------------------------------------------------------- #
# Pasted text -> grid
# --------------------------------------------------------------------------- #
def grid_from_text(text: str) -> list:
    """Pull X links out of arbitrarily messy pasted text.

    The frozen loader's plain-list mode takes the WHOLE CELL containing a URL as
    the link, which is right for a tidy one-link-per-line paste and wrong for
    "check this out https://x.com/a/status/1 lol" — that would become the link.
    So the URLs are extracted here first and handed over one per row, which is a
    shape `_rows_from_grid` reads exactly as intended.

    Handles: several links on one line, junk words around them, blank lines,
    bullets and numbering, trailing punctuation, and angle-bracket wrapping.
    """
    rows = []
    for line in (text or "").splitlines():
        for raw in _URL_RE.findall(line):
            cleaned = input_loader._clean_url(raw.strip("<>“”‘’"))
            if cleaned:
                rows.append([cleaned])
    return rows


def grid_from_csv_text(text: str) -> list:
    """CSV text (from a fetched sheet) -> grid, using the same delimiter
    sniffing an uploaded .csv gets, so a sheet and a downloaded copy of that
    sheet parse identically."""
    if not (text or "").strip():
        return []
    delim = _sniff_delimiter(text, ",")
    if not isinstance(delim, str) or len(delim) != 1:
        delim = ","
    try:
        rows = list(csv.reader(io.StringIO(text), delimiter=delim))
    except csv.Error:
        rows = [[line] for line in text.splitlines()]
    return [[(c or "").strip() for c in row] for row in rows]


def detect_columns(grid: list) -> dict:
    """Which column holds the links, which holds the account, and what else is
    on offer.

    `input_loader._header_index` scans only the FIRST FIVE rows and takes the
    first column whose header matches — silently, with no way to correct it.
    This exposes the same decision so the user can see it and override it.
    """
    header = input_loader._header_index(grid)
    header_row = header[0] if header else None
    cmap = header[1] if header else {}
    guessed = False

    # The frozen loader only recognises a header row when one of its column
    # names matches its own list — "Post URL" is not in it, for instance. When
    # it finds none, a first row of plain text is still what a human reads as
    # the headings, so it is used for LABELS (and as the body start for a
    # reshape). Parsing is untouched: this never changes what the loader does.
    if header_row is None and grid:
        first = grid[0]
        text_cells = [c for c in first if c and not _URL_RE.search(c)]
        if len(text_cells) >= 2 and not any(_URL_RE.search(c or "") for c in first):
            header_row, guessed = 0, True

    body = grid[(header_row + 1):] if header_row is not None else grid

    width = max((len(r) for r in grid), default=0)
    columns = []
    for i in range(width):
        name = ""
        if header_row is not None and i < len(grid[header_row]):
            name = grid[header_row][i]
        sample = ""
        for row in body:
            if i < len(row) and row[i]:
                sample = row[i]
                break
        role = None
        if cmap.get("link") == i:
            role = "link"
        elif cmap.get("account") == i:
            role = "account"
        elif cmap.get("category") == i:
            role = "category"
        elif not header and any(
                i < len(r) and _URL_RE.search(r[i] or "") for r in body[:40]):
            role = "link"        # plain list: the column that holds URLs
        columns.append({
            "index": i,
            "name": name or f"Column {chr(65 + i) if i < 26 else i + 1}",
            "sample": (sample or "")[:90],
            "role": role,
        })
    return {"has_header": bool(header), "guessed_header": guessed,
            "header_row": header_row, "columns": columns}


def reshape(grid: list, link_col=None, account_col=None) -> list:
    """Rebuild a grid as a canonical Account | Link table.

    Those exact headers are what `input_loader._header_index` recognises, so the
    frozen loader reads back precisely the columns the user chose. Returns the
    grid untouched when no override was asked for.
    """
    if link_col is None:
        return grid
    try:
        link_i = int(link_col)
    except (TypeError, ValueError):
        raise UploadError("The chosen link column is not valid.")
    acc_i = None
    if account_col not in (None, "", -1, "-1"):
        try:
            acc_i = int(account_col)
        except (TypeError, ValueError):
            acc_i = None

    info = detect_columns(grid)
    start = (info["header_row"] + 1) if info["header_row"] is not None else 0
    out = [["Account", "Link"]]
    for row in grid[start:]:
        link = row[link_i] if link_i < len(row) else ""
        if not link:
            continue
        account = row[acc_i] if acc_i is not None and acc_i < len(row) else ""
        out.append([account, link])
    if len(out) == 1:
        raise UploadError("That column has no values in it — pick another.")
    return out


# --------------------------------------------------------------------------- #
# Grid -> rows + diagnostics
# --------------------------------------------------------------------------- #
def _status_key(link: str) -> str:
    """Identity of a post for duplicate detection: its status id when present,
    else the URL minus tracking query junk. `?s=20&t=...` makes the same post
    look like several different ones."""
    m = re.search(r"/status/(\d+)", link or "")
    if m:
        return m.group(1)
    return re.sub(r"[?#].*$", "", (link or "").lower()).rstrip("/")


def analyse(grid: list, dedupe: bool = False, platform: str = "x") -> dict:
    """Rows the pipeline would capture, plus WHY anything was left out.

    The frozen `_rows_from_grid` silently drops non-X links and reports only a
    count. Row numbers are computed here, alongside it, so a 200-row sheet says
    "row 14" instead of starting a manual hunt (roadmap A5). The loader itself
    is untouched and remains the single source of truth for what a sheet means.
    """
    import netlinks                     # profiles/netlinks.py (on sys.path)
    rows = netlinks.rows_from_grid(grid, platform)
    is_ours = netlinks.MATCHERS.get(platform, input_loader.is_x_url)
    ours = {"x": "an x.com / twitter.com", "facebook": "a facebook.com",
            "instagram": "an instagram.com",
            "combined": "an X / Facebook / Instagram"}.get(platform, f"a {platform}")

    kept = set()
    for r in rows:
        kept.add(r["link"])

    dropped = []
    seen_raw = set()
    for line_no, cells in enumerate(grid, start=1):
        for cell in cells:
            for raw in _URL_RE.findall(cell or ""):
                cleaned = input_loader._clean_url(raw)
                # Compare what the READER would have kept, not the raw text: it
                # normalises a Facebook/Instagram link (host, tracking query),
                # so a captured row would otherwise be listed as rejected right
                # under itself.
                canonical = netlinks.normalize_url(cleaned, platform)
                if canonical in kept or canonical in seen_raw:
                    continue
                seen_raw.add(canonical)
                reason = (f"this is not {ours} link"
                          if not is_ours(cleaned)
                          else f"not recognised as a {platform} post link")
                dropped.append({"row": line_no, "value": cleaned[:180],
                                "reason": reason})

    # Duplicates are REPORTED always and removed only on request: the frozen
    # loader keeps them, and silently changing that would alter what an existing
    # sheet produces.
    positions = {}
    for i, r in enumerate(rows, start=1):
        positions.setdefault(_status_key(r["link"]), []).append(i)
    duplicates = [{"link": rows[v[0] - 1]["link"], "positions": v}
                  for v in positions.values() if len(v) > 1]

    if dedupe and duplicates:
        first = {v[0] for v in positions.values()}
        rows = [r for i, r in enumerate(rows, start=1) if i in first]

    return {
        "rows": rows,
        "dropped": dropped,
        "duplicates": duplicates,
        "duplicate_count": sum(len(d["positions"]) - 1 for d in duplicates),
        "over_limit": len(rows) > config.MAX_LINKS,
        "limit": config.MAX_LINKS,
    }


def suffix_of(filename: str) -> str:
    suffix = Path((filename or "").replace("\\", "/")).suffix.lower()
    if suffix not in ALLOWED_SUFFIXES:
        raise UploadError(
            f"'{safe_upload_name(filename)}' is not a supported file type. "
            "Accepted: " + ", ".join(ALLOWED_SUFFIXES) + ".")
    return suffix

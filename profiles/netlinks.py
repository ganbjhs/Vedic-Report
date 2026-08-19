"""Which network a link belongs to, and a platform-neutral row reader.

The frozen `src/input_loader._rows_from_grid` is X-only by construction (its
last line drops every non-X link). A second platform needs the SAME sheet
semantics — header detection, the plain-list mode where a URL-less line is a
category, comment lines, sibling-cell account names — minus that one filter.
So the layout logic is reused **read-only** (`_header_index`, `_clean_url`,
`_URL_RE`, `_IGNORE_HEADERS` are imported, never copied or edited) and only the
keep-test is parameterised.

Used by both `prof_runner` (what the job captures) and `webapp/uploads.analyse`
(what the preview shows), so the two cannot disagree about what a Facebook link
is. Nothing here imports Playwright.

It also reads the **sectioned** sheet the team keeps by hand: no link column and
no handle column, section names in column A, and the header row's own first cell
naming the first section. That is recognised from its shape (`metric_header`),
not from a mode the user has to pick — see RULEBOOK §18c.
"""
import re
import sys
from pathlib import Path

_SRC = Path(__file__).resolve().parents[1] / "src"
if str(_SRC) not in sys.path:
    sys.path.insert(0, str(_SRC))
import input_loader          # noqa: E402  frozen, read-only

_FB_HOSTS = ("facebook.com", "fb.com", "fb.watch", "m.facebook.com",
             "mbasic.facebook.com", "web.facebook.com")

# What a Facebook *post* URL looks like. Anything on facebook.com that is not
# one of these (a Page home, /marketplace, /groups/<id> with no post) is not
# something a single-post capture can frame.
_FB_POST_RE = re.compile(
    r"facebook\.com/(?:"
    r"[^/?#]+/posts/[^/?#]+"            # /<page>/posts/<id or pfbid>
    r"|[^/?#]+/photos/[^/?#]+/\d+"      # /<page>/photos/a.xxx/<id>
    r"|photo(?:\.php)?/?\?.*fbid=\d+"   # /photo/?fbid=…  /photo.php?fbid=…
    r"|permalink\.php\?.*story_fbid="   # /permalink.php?story_fbid=…&id=…
    r"|story\.php\?.*story_fbid="       # mobile permalink
    r"|[^/?#]+/videos/[^/?#]+"          # /<page>/videos/<id>
    r"|watch/?\?.*v=\d+"                # /watch/?v=…
    r"|reel/\d+"                        # /reel/<id>
    r"|share/[pv]/[^/?#]+"              # /share/p/<code>  share links
    r"|groups/[^/?#]+/(?:posts|permalink)/[^/?#]+"
    r")", re.I)


def is_x_url(url: str) -> bool:
    return input_loader.is_x_url(url)


def is_fb_url(url: str) -> bool:
    u = (url or "").lower()
    return any(h in u for h in _FB_HOSTS) or "fb.watch/" in u


def is_fb_post_url(url: str) -> bool:
    u = (url or "").lower()
    return "fb.watch/" in u or bool(_FB_POST_RE.search(u))


def normalize_fb_url(url: str) -> str:
    """Desktop host, no tracking junk. m./mbasic./web. all serve the same post;
    the desktop layout is the one the capture is written against."""
    u = input_loader._clean_url(url)
    u = re.sub(r"^https?://(m|mbasic|web)\.facebook\.com", "https://www.facebook.com", u, flags=re.I)
    u = re.sub(r"^https?://facebook\.com", "https://www.facebook.com", u, flags=re.I)
    if "?" in u and "fbid=" not in u and "story_fbid=" not in u and "v=" not in u:
        # /posts/<id>?__cft__=… → /posts/<id>; but a COMMENT permalink keeps
        # its comment_id — that is the thing the capture frames (v3).
        base, q = u.split("?", 1)
        m = re.search(r"(?:^|&)(comment_id=\d+)", q)
        u = base + (f"?{m.group(1)}" if m else "")
    return u


_IG_POST_RE = re.compile(r"instagram\.com/(?:[^/?#]+/)?(?:p|reel|reels|tv)/[A-Za-z0-9_-]+", re.I)


def is_ig_url(url: str) -> bool:
    return "instagram.com" in (url or "").lower()


def is_ig_post_url(url: str) -> bool:
    return bool(_IG_POST_RE.search((url or "").lower()))


def normalize_ig_url(url: str) -> str:
    u = input_loader._clean_url(url)
    u = re.sub(r"^https?://(m\.)?instagram\.com", "https://www.instagram.com", u, flags=re.I)
    u = re.sub(r"^https?://instagram\.com", "https://www.instagram.com", u, flags=re.I)
    u = u.split("?", 1)[0]
    return u if u.endswith("/") else u + "/"


def normalize_url(url: str, platform: str = "combined") -> str:
    """The link exactly as `rows_from_grid` would keep it.

    The preview needs this: it walks the raw grid looking for URLs that did NOT
    survive, and an Instagram link only fails that test because the reader
    dropped its `?igsh=…`. Comparing raw text would report the row the user can
    see in the table right above as rejected."""
    link = input_loader._clean_url(url)
    plat = platform_of(link) if platform == "combined" else platform
    if plat == "facebook":
        return normalize_fb_url(link)
    if plat == "instagram":
        return normalize_ig_url(link)
    return link


def platform_of(url: str):
    """'x' | 'facebook' | 'instagram' | None for a URL."""
    if is_x_url(url):
        return "x"
    if is_fb_url(url):
        return "facebook"
    if is_ig_url(url):
        return "instagram"
    return None


def is_any_url(url: str) -> bool:
    return platform_of(url) is not None


MATCHERS = {"x": is_x_url, "facebook": is_fb_url, "instagram": is_ig_url,
            "combined": is_any_url}

# One column, two metrics. "Reach/views" is how the team's own sheet heads the
# single number it has for both, so that column feeds `views` AND `reach` with
# the same value rather than forcing a choice the sheet did not make.
_REACH_VIEWS = ("reach/views", "views/reach", "reach / views", "views / reach")

# Extra sheet columns a combined report may carry. Header text -> metric key.
# Values are printed as typed (the team reads them from Insights); nothing here
# is scraped. Header match is case-insensitive, punctuation-insensitive.
METRIC_HEADERS = {
    "like": ("like", "likes", "reactions", "reaction"),
    "impressions": ("impression", "impressions", "post impression", "post impressions"),
    "views": ("views", "video views", "view", "plays") + _REACH_VIEWS,
    "reach": ("reach", "post reach") + _REACH_VIEWS,
    "comments": ("comments", "comment"),
    "shares": ("shares", "share", "reposts", "retweets"),
    "followers": ("followers",),
}
_SECTION_HEADERS = ("section", "category", "group", "type", "bucket")
_HANDLE_HEADERS = ("handle", "handle name", "account", "account name", "page name",
                   "name", "author", "page", "profile")
# A header cell that names the link column. `input_loader._LINK_HEADERS` is the
# frozen list; these are its normalised forms plus the spellings this reader has
# seen in the wild.
_LINK_HEADER_CELLS = tuple(sorted(
    {re.sub(r"[^a-z0-9 ]+", "", h) for h in input_loader._LINK_HEADERS}
    | {"post url", "posturl"}))


def _norm(h: str) -> str:
    """Lower-case, punctuation-free, single-spaced — so 'Post Impression',
    'post_impression' and 'Post  Impression' are one header, and 'Reach/views'
    normalises the same way whichever slash spacing the sheet used."""
    return re.sub(r"\s+", " ",
                  re.sub(r"[^a-z0-9 ]+", "", (h or "").lower())).strip()


_METRIC_NAMES = {k: {_norm(n) for n in names} for k, names in METRIC_HEADERS.items()}


def _has_letters(s: str) -> bool:
    """True if `s` holds at least one letter. A cell of digits, commas and
    spaces ('676  63,000') is a stray metric value, never a section name."""
    return bool(re.search(r"[^\W\d_]", s or "", re.UNICODE))


def metric_header(grid):
    """`(row_index, {metric_key: column_index}, first_section_or_None)` for the
    sheet's header row, or None.

    Two shapes are recognised, and neither needs a mode switch.

    * The ordinary one names a **link column** and puts the metric headers
      beside it; there is no section in the header row, so the third item is
      None.
    * The **sectioned** one — what the team's own sheet looks like — has no link
      header at all. Column A holds the FIRST section's name and the cells
      beside it are the metric headers; every later section is a plain row with
      its name in column A. Two or more metric headers sitting next to a piece
      of ordinary text is the whole signal.
    """
    for i, cells in enumerate(grid[:5]):
        low = [_norm(c) for c in cells]
        found = {}
        for key, names in _METRIC_NAMES.items():
            j = next((j for j, c in enumerate(low) if c in names), None)
            if j is not None:
                found[key] = j
        if not found:
            continue
        if any(c in _LINK_HEADER_CELLS for c in low):
            return i, found, None
        first = cells[0] if cells else ""
        if 0 in found.values() or input_loader._URL_RE.search(first):
            continue        # column A is itself a metric header, or a link
        if len({j for j in found.values() if j > 0}) >= 2 and _has_letters(first):
            return i, found, first.strip()
    return None


def metric_columns(grid) -> dict:
    """{metric_key: column_index} from the header row, if any — including a
    sectioned sheet's header, whose first cell is a section name."""
    head = metric_header(grid)
    return head[1] if head else {}


def fb_name(url: str) -> str:
    """'<page>' from /<page>/posts/… when present, else 'Facebook post'."""
    m = re.search(r"facebook\.com/([^/?#]+)/(?:posts|photos|videos)/", (url or "").lower())
    if m and m.group(1) not in ("photo", "watch", "reel", "share", "groups", "permalink.php"):
        return m.group(1)
    return "Facebook post"


def _display_name(link: str, plat: str) -> str:
    if plat == "x":
        return input_loader.derive_name(link)
    if plat == "facebook":
        return fb_name(link)
    m = re.search(r"instagram\.com/([^/?#]+)/(?:p|reel|reels|tv)/", (link or "").lower())
    return ("@" + m.group(1)) if m else "Instagram post"


def rows_from_grid(grid, platform: str = "x") -> list:
    """The frozen reader's rows for X; the same layout rules for any other
    platform, keeping that platform's links instead. Rows carry the platform
    (per row for 'combined'), and any metric columns the sheet has."""
    if platform == "x":
        return input_loader._rows_from_grid(grid)
    keep = MATCHERS[platform]
    grid = [cells for cells in grid if any(cells)]
    if not grid:
        return []
    header = input_loader._header_index(grid)
    mcols = metric_columns(grid)
    rows = []

    def row(category, account, link, cells=()):
        link = input_loader._clean_url(link)
        plat = platform_of(link) if platform == "combined" else platform
        link = normalize_url(link, platform)
        r = {"category": (category or "Uncategorized").strip() or "Uncategorized",
             "account_name": (account or "").strip() or _display_name(link, plat or "x"),
             "link": link, "post_link": link, "platform": plat or platform}
        if not (account or "").strip():
            # The sheet has no handle column, so `account_name` above is a
            # placeholder derived from the URL — and for /61559.../posts/ or
            # /share/p/ it is a page id or nothing at all. Flag it so the worker
            # can put the name the capture actually read in its place.
            r["account_auto"] = True
        metrics = {}
        for key, j in mcols.items():
            if j < len(cells) and str(cells[j]).strip():
                metrics[key] = str(cells[j]).strip()
        if metrics:
            r["sheet_metrics"] = metrics
        return r

    if header:
        start, cmap = header
        for cells in grid[start + 1:]:
            link = cells[cmap["link"]] if cmap["link"] < len(cells) else ""
            if not input_loader._URL_RE.search(link):
                link = next((c for c in cells if input_loader._URL_RE.search(c)), "")
            if not link:
                continue
            account = cells[cmap["account"]] if cmap["account"] is not None \
                and cmap["account"] < len(cells) else ""
            category = cells[cmap["category"]] if cmap["category"] is not None \
                and cmap["category"] < len(cells) else ""
            rows.append(row(category, account, link, cells))
    else:
        # Plain-list mode, and the sectioned sheet's mode: column A carries
        # either a post URL or a section name. A header row (whether it names a
        # link column or only metrics) is never a section — the body starts
        # after it — and its first cell is the first section when the sheet is
        # sectioned.
        head = metric_header(grid)
        start = (head[0] + 1) if head else 0
        category = (head[2] if head else None) or "Uncategorized"
        mcol_idx = set(mcols.values())
        for cells in grid[start:]:
            if cells and cells[0].lstrip().startswith("#"):
                continue
            first = cells[0] if cells else ""
            if not input_loader._URL_RE.search(first) and not first.strip():
                # Column A empty. A row like ["", "676", "63,000", "63,000"]
                # sits under every block of the team's sheet; it is neither a
                # post nor a section, so it is dropped rather than promoted to
                # a category made of numbers.
                continue
            url = first if input_loader._URL_RE.search(first) else \
                next((c for c in cells if input_loader._URL_RE.search(c)), "")
            if url:
                account = next((c for j, c in enumerate(cells)
                                if c and j not in mcol_idx
                                and not input_loader._URL_RE.search(c)
                                and not re.fullmatch(r"[\d,]+", c)), "")
                rows.append(row(category, account, url, cells))
                continue
            joined = " ".join(c for j, c in enumerate(cells)
                              if c and j not in mcol_idx)
            if joined.lower() in input_loader._IGNORE_HEADERS:
                continue
            if not _has_letters(first):
                continue        # digits / commas / spaces only — not a section
            category = joined.strip() or category

    kept = [r for r in rows if keep(r["link"])]
    dropped = len(rows) - len(kept)
    if dropped:
        print(f"[input] skipped {dropped} non-{platform} link(s)", flush=True)
    return kept


def load_rows(source: str, platform: str = "x") -> list:
    """Rows from an .xlsx (or a plain text list) for `platform`. X goes through
    the frozen loader unchanged; other platforms read the same file shapes."""
    if platform == "x":
        return input_loader.load(source)
    path = Path(source)
    print(f"[input] reading {source}", flush=True)
    grid = None
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True, data_only=True)
            ws = wb.active
            grid = [[("" if c is None else str(c)).strip() for c in r]
                    for r in ws.iter_rows(values_only=True)]
            wb.close()
        except Exception:
            grid = None
    if grid is None:
        text = path.read_text(encoding="utf-8", errors="replace")
        grid = [[c.strip() for c in re.split(r"[\t,]", line)] for line in text.splitlines()]
    return rows_from_grid(grid, platform)
